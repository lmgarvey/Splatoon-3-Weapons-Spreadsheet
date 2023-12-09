# L.M. Garvey, 2023

# program to create a splatoon 3 weapons spreadsheet
# uses Workbook from openpyxl
# the spreadsheet will be stylized to all lowercase, and saved in the same directory as this program
# https://github.com/lmgarvey/Splatoon-3-Weapons-Spreadsheet

# takes as input from user: 'current points' for a prompted weapon
#       you can check this number by clicking ZR twice with the
#       specified weapon selected in-game and looking at the
#       number in the bar graph (your 'freshness' for that weapon).
#       each weapon will be prompted individually. as of december 2023,
#       there are 110 weapons.
# using this, produces a spreadsheet with several sheets:
#       1. the `numbers` sheet, including progress toward the next
#          freshness, as well as several default values, such as
#          each weapon's sub and special, range, and more.
#       2. the basic chart calculations, which performs calculations
#          based on the user's inputs to the `numbers` sheet to create
#          pie charts showing where they've earned most of their points,
#          in terms of weapon class, subs, specials, and more.
#       3. the basic pie charts, based on the calculations made in the
#          directly previous sheet.
#       4. chart calculations, but only for weapons with less than five
#          stars. there are 1 million points between four and five stars,
#          so it is valuable to remove those outliers.
#       5. pie charts for weapons with less than five stars, again based
#          on the calculations made in the directly previous sheet.
# for the class-specific values, such as impact, damage, ink speed,
#          charge speed, the value will be in its appropriate column, and
#          the corresponding chart will be in the cell directly to the right.
#          to keep this consistent, column W is titled 'mobility chart' to
#          hold the charts for the mobility numbers.

# future plans/ideas:
# automate creating calculations and charts for less than
#       four stars, <3, <2, <1, by generalizing the <5 process
# calculations and charts for non-unique things, eg shooter ranges vs charger ranges
#       and which subs are more common to certain classes
# matchmaking numbers across classes


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils.exceptions import InvalidFileException

import build_weapons
import update_existing


def set_up_weapons() -> dict:
    """
    Sets up all the weapons alphabetically within their own classes
    Defaults each weapon to have zero points
    Calls build_[class] for each class in build_weapons.py
    :return: dict of
                    {class:
                        {
                        spec A: ( name of class specific A, offset from first spec A column as 0-3 )
                        spec B: ( name of class specific B, offset from first spec B column as 0-3 )
                        weapon_in_class:
                            {
                            sub: name of sub
                            special: name of special
                            points: current freshness points
                            range: range as a number
                            role: role(s)
                            weight: light, middle, or heavy
                            class_specific_1: value as number
                            class_specific_2: value as number
                            special points: points needed for special
                            }
                        }
                    }
    """

    blaster_dct = build_weapons.build_blasters()
    brella_dct = build_weapons.build_brellas()
    brush_dct = build_weapons.build_brushes()
    charger_dct = build_weapons.build_chargers()
    dualies_dct = build_weapons.build_dualies()
    roller_dct = build_weapons.build_rollers()
    shooter_dct = build_weapons.build_shooters()
    slosher_dct = build_weapons.build_sloshers()
    splatana_dct = build_weapons.build_splatanas()
    splatling_dct = build_weapons.build_splatlings()
    stringer_dct = build_weapons.build_stringers()

    out = {"blaster": blaster_dct, "brella": brella_dct, "brush": brush_dct, "charger": charger_dct,
           "dualies": dualies_dct, "roller": roller_dct, "shooter": shooter_dct, "slosher": slosher_dct,
           "splatana": splatana_dct, "splatling": splatling_dct, "stringer": stringer_dct}
    return out


def collect_points(dct: dict, default_zeros=False) -> None:
    """
    Outer function that calls collect_points for each class in dct
    :param dct: {class:
                    {weapon_in_class: points}
                 }
    :param default_zeros:
        True: fill all weapon points with zero
        False: prompt user for their points
    :return: None
    """
    for weapon_class in dct:
        if not default_zeros:
            print("\n")
            print("{} class!".format(weapon_class))

        for weapon in dct[weapon_class]:
            if weapon == "spec A" or weapon == "spec B":
                continue
            if not default_zeros:
                print("Current points for '{}', or 'q' to quit and default to zeros:".format(weapon))
                points = input()
                if points.upper() == "Q":
                    default_zeros = True
                    points = 0
                    print("Stopping points collection, filling in remaining weapons with zero points.\n")
                else:
                    try:
                        points = int(points)
                    except ValueError:
                        print("Invalid points input, settings points for {} = 0.\n".format(weapon))
                        points = 0
            else:
                points = 0
            if not points or points < 0:
                points = 0
            if points > 1160000:
                points = 1160000
            dct[weapon_class][weapon]["points"] = points

        if not default_zeros:
            print("~~~~~~~~~~")


def set_header_row(sheet) -> None:
    """
    Sets up the header row
    :param sheet: sheet to set the header row on
    :return: None
    """
    sheet["A1"] = "class"
    sheet["B1"] = "weapon"
    sheet["C1"] = "sub"
    sheet["D1"] = "special"
    sheet["E1"] = "freshness"
    sheet["F1"] = "current"
    sheet["G1"] = "checkpoint"
    sheet["H1"] = "necessary"
    sheet["I1"] = "progress (%)"
    sheet["J1"] = "progress (chart)"
    sheet["K1"] = "range (#)"
    sheet["L1"] = "range (chart)"
    sheet["M1"] = "role"
    sheet["N1"] = "weight/speed"
    sheet["O1"] = "impact"
    sheet["P1"] = "damage"
    sheet["Q1"] = "ink speed"
    sheet["R1"] = "charge speed"
    sheet["S1"] = "fire rate"
    sheet["T1"] = "durability"
    sheet["U1"] = "handling"
    sheet["V1"] = "mobility"
    sheet["W1"] = "special points"


def set_up_sheet(sheet, weapons) -> None:
    """
    Sets up basic outline of sheet - column headers,
        class names, weapon names, subs, specials,
        default zero points, range, role, weight,
        class-spec-A, class-spec-B, points for special
    :param sheet: sheet to set up
    :param weapons: {class:
                        {
                        spec A: (name of class spec A, offset from first spec A column)
                        spec B: (name of class spec B, offset from first spec B column)
                        weapon: dict of attribute : value
                        }
                     }
    :return: None
    """
    set_header_row(sheet)

    # set up class and weapon names in columns A and B
    start = 2
    for weapon_class in weapons:

        # for getting each weapon into its own cell
        weapons_dct_as_lst = list()
        for weapon in weapons[weapon_class]:
            if weapon != "spec A" and weapon != "spec B":
                weapons_dct_as_lst.append(weapon)

        class_spec_1 = weapons[weapon_class]["spec A"][0]
        class_spec_2 = weapons[weapon_class]["spec B"][0]
        spec_1_offset = weapons[weapon_class]["spec A"][1]
        spec_2_offset = weapons[weapon_class]["spec B"][1]

        end = len(weapons[weapon_class]) + start - 2  # -2 discounts the spec_A and spec_B items
        for row in range(start, end):
            # set weapon class in column A
            sheet.cell(row, 1).value = weapon_class

            weapon = weapons_dct_as_lst[row - start]
            sheet.cell(row, 2).value = weapon  # column B
            sheet.cell(row, 3).value = weapons[weapon_class][weapon]["sub"]  # C
            sheet.cell(row, 4).value = weapons[weapon_class][weapon]["special"]  # D
            sheet.cell(row, 6).value = weapons[weapon_class][weapon]["points"]  # F
            sheet.cell(row, 11).value = weapons[weapon_class][weapon]["range"]  # K
            sheet.cell(row, 13).value = weapons[weapon_class][weapon]["role"]  # M
            sheet.cell(row, 14).value = weapons[weapon_class][weapon]["weight"]  # N
            sheet.cell(row, 15 + spec_1_offset).value = weapons[weapon_class][weapon][class_spec_1]  # O, P, Q, R
            sheet.cell(row, 19 + spec_2_offset).value = weapons[weapon_class][weapon][class_spec_2]  # S, T, U, V
            sheet.cell(row, 23).value = weapons[weapon_class][weapon]["special points"]  # W

        start = end

    # set up range chart number in column directly after range-as-number
    range_rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                             color="000000", showValue=False)
    for row in range(2, 112):
        sheet.cell(row, 12).value = "=K{}".format(row)
    sheet.conditional_formatting.add('L2:L111', range_rule)

    # range chart in its own column (L), class specs in same column as value (O, P, .., U, V)
    charts_to_add = ['O2:O111', 'P2:P111', 'Q2:Q111', 'R2:R111',
                     'S2:S111', 'T2:T111', 'U2:U111', 'V2:V111']
    specs_rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                             color="000000", showValue=True)
    for column in charts_to_add:
        sheet.conditional_formatting.add(column, specs_rule)

    # weapon class, weapon name, and column header always visible:
    sheet.freeze_panes = "C2"

    # header row in bold
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # match width to widest cell in column
    # https://stackoverflow.com/a/35790441
    dims = dict()
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, val in dims.items():
        sheet.column_dimensions[col].width = val + 3


def set_up_points_columns(sheet) -> None:
    """
    Modifies points-related columns:
        E: freshness in stars
        G: next freshness checkpoint
        H: points remaining to next checkpoint
        I: progress to next checkpoint as a percentage
        J: progress to next checkpoint as a bar chart
    User's points are held in column F
    :param sheet: sheet to modify
    :return: None
    """

    for row in range(2, 112):
        # E: freshness in stars
        sheet.cell(row, 5).value = "=IF(F{} < 10000, \"☆\", " \
                                   "IF(F{} < 25000, \"★\", " \
                                   "IF(F{} < 60000, \"★★\", " \
                                   "IF(F{} < 160000, \"★★★\", " \
                                   "IF(F{} < 1160000, \"★★★★\", \"★★★★★\")))))" \
            .format(row, row, row, row, row)

        # G: next freshness checkpoint
        sheet.cell(row, 7).value = "=IF(F{} < 10000, 10000, " \
                                   "IF(F{} < 25000, 25000, " \
                                   "IF(F{} < 60000, 60000, " \
                                   "IF(F{} < 160000, 160000, " \
                                   "IF(F{} < 1160000, 1160000, 1160000)))))" \
            .format(row, row, row, row, row)

        # H: points remaining to next checkpoint
        sheet.cell(row, 8).value = "=G{} - F{}".format(row, row)

        # I: progress to next checkpoint as a percentage, will be 'X' if 5-starred
        sheet.cell(row, 9).value = "=IF(H{} = 0, \"X\", F{} * 100 / G{})".format(row, row, row)

        # for 5-star weapons, progress (%) shows an 'X', get that right-aligned
        sheet.cell(row, 9).alignment = Alignment(horizontal='right')

        # round progress percentage to two decimals
        sheet.cell(row, 9).number_format = '#,##0.00'

        # J: progress as bar chart, need the numbers in there first...
        sheet.cell(row, 10).value = "=IF(H{} = 0, \"\", F{} * 100 / G{})".format(row, row, row)

    # ...J again: progress bar, hide the numbers, replace them with a data bar
    rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                       color="000000", showValue=False)
    sheet.conditional_formatting.add('J2:J111', rule)


def create_new_sheet() -> None:
    """
    Creates a new spreadsheet
    :return: None
    """
    book = Workbook()
    numbers_sheet = book.active
    numbers_sheet.title = "numbers"

    weapons_dct = set_up_weapons()
    collect_points(weapons_dct, default_to_zeros)
    set_up_sheet(numbers_sheet, weapons_dct)
    set_up_points_columns(numbers_sheet)

    book.save(filename="splatoon_weapons.xlsx")

    print()
    print("Done! Your spreadsheet is ready for you in the same directory as this python file.")

    formulas_sheet = book.create_sheet("formulas")


if __name__ == '__main__':
    print("Welcome to `Creating your Splatoon 3 weapons spreadsheet!`\n")
    print()
    print("This program will prompt you with one weapon at a time, and ask for your current points on that weapon.\n")
    print("Type the relevant number when prompted, then press `enter`. Repeat this for all 101 weapons.\n"
          "If you do not have the prompted weapon, you can type `0`, or just press `enter` with no input.\n"
          "The weapons will be prompted by class, and alphabetically within that class.\n"
          "If at any time you wish to stop, type 'q', and all remaining weapons will be defaulted to zero points.\n")
    print()
    print("At this time, you can choose to (A) fill in your points, or (B) have a spreadsheet created \n"
          "with all of the point values defaulted to zero. If you select (A), you may also type `q` at any\n"
          "time to stop, and have the remaining weapons all defaulted to zero.\n"
          "Please enter your selection:\n"
          " A: Create sheet and fill in your own point values\n"
          " B: Create sheet and default all points to zero\n"
          " C: Update an existing sheet\n")
    choice = input()
    while choice.upper() != "A" and choice.upper() != "B" and choice.upper() != "C":
        print("Sorry, please make sure you type only A, B, or C.\n"
              "A: Fill in your own point values\n"
              "B: Default all points to zero\n"
              "C: Update an existing sheet\n")
        choice = input()

    default_to_zeros = False
    if choice.upper() == "B":
        default_to_zeros = True

    loading = False
    if choice.upper() == "C":
        print("To update an existing sheet, please make sure it has the \'.xlsx\' extension"
              "type, and that it is in the same directory as this python file.\n"
              "Please enter the name of the existing sheet, *exactly* as it appears in the directory.\n")
        choice = input()
        while True:
            try:
                workbook = load_workbook(choice)
                loading = True
                break
            except InvalidFileException:
                print("Invalid file name. Please enter one of the following:\n"
                      "A: Create a new sheet and fill in your own point values\n"
                      "B: Create a new sheet and default all points to zero\n"
                      "C: Enter the file name of an existing sheet, as \"example.xlsx\"\n")
                choice = input()
                if choice.upper() == "A":
                    # proceed as normal with collecting points
                    break
                elif choice.upper() == "B":
                    # create new sheet, default to zeros
                    default_to_zeros = True
                    break

    if loading:
        # option C, loading into an existing worksheet
        print("Loading into existing worksheet named: {}\n"
              "Note: the program assumes that your weapon names are stored in column B,\n"
              "and your current points are stored in column F.\n".format(choice))
        book = load_workbook(choice)
        update_existing.update_wrapper(book, choice)

    else:
        create_new_sheet()
